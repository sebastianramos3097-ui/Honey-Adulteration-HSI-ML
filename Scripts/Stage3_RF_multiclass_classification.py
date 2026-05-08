# -*- coding: utf-8 -*-
"""
Stage 3 (ML) — Random Forest multiclass classification (RandomizedSearchCV)
Target: concentration_class (e.g., 0, 5, 10, 25, 50, 100)
Split: Train = Acquisition 1–5, Test = Acquisition 6 (external test set)

IMPORTANT:
- Spectral bands are already z-score standardized in the dataset documentation.
  We do NOT apply any additional scaling. We only report diagnostics.

Outputs:
- Excel workbook with data checks, tuning results, best parameters, test metrics, confusion matrix, classification report
- Confusion matrix figure (PNG)
- Text log file

Designed to run in Visual Studio Code.
"""

import os
import sys
import time
import json
import logging

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import StratifiedKFold, RandomizedSearchCV
from sklearn.metrics import (
    accuracy_score,
    f1_score,
    classification_report,
    confusion_matrix,
    balanced_accuracy_score
)

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment

# Random distributions (no SciPy needed)
# We will sample integers using numpy RandomState inside PARAM_DIST definitions.


# -----------------------------
# CONFIG
# -----------------------------
DATA_PATH = r"C:\Sebastian\Pythonsebastian\adulteration_dataset_26_08_2021.xlsx"

RANDOM_STATE = 42
N_SPLITS_CV = 5

# RandomizedSearchCV settings
N_ITER = 40  # 40 * 5-fold = 200 fitted RF models (fast enough on a laptop)

# Output folder structure (relative to this script location)
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
STEP_DIR = os.path.join(SCRIPT_DIR, "Step3_ML_RF_Classification_Randomized")
OUTPUTS_DIR = os.path.join(STEP_DIR, "outputs")
FIGURES_DIR = os.path.join(STEP_DIR, "figures")
LOGS_DIR = os.path.join(STEP_DIR, "logs")

EXCEL_OUT = os.path.join(OUTPUTS_DIR, "RF_multiclass_results_randomized.xlsx")
CONF_MATRIX_FIG = os.path.join(FIGURES_DIR, "confusion_matrix_test.png")
LOG_FILE = os.path.join(LOGS_DIR, "RF_multiclass_run_log.txt")

SCORING = "f1_macro"


# -----------------------------
# HELPERS
# -----------------------------
def setup_folders():
    for d in [STEP_DIR, OUTPUTS_DIR, FIGURES_DIR, LOGS_DIR]:
        os.makedirs(d, exist_ok=True)


def setup_logger():
    logger = logging.getLogger("RF_Multiclass_Randomized")
    logger.setLevel(logging.INFO)
    logger.handlers = []

    fmt = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")

    fh = logging.FileHandler(LOG_FILE, mode="w", encoding="utf-8")
    fh.setFormatter(fmt)
    logger.addHandler(fh)

    sh = logging.StreamHandler(sys.stdout)
    sh.setFormatter(fmt)
    logger.addHandler(sh)

    return logger


def read_excel_dataset(path: str) -> pd.DataFrame:
    return pd.read_excel(path, engine="openpyxl")


def normalize_str(s: str) -> str:
    return "".join(ch for ch in str(s).strip().lower() if ch.isalnum())


def find_column(df: pd.DataFrame, candidates: list) -> str | None:
    """
    Find a column whose normalized name matches candidates (robust to case/spaces/underscores).
    Also tries substring matches if exact match fails.
    """
    norm_map = {normalize_str(c): c for c in df.columns}
    cand_norms = [normalize_str(x) for x in candidates]

    for cn in cand_norms:
        if cn in norm_map:
            return norm_map[cn]

    # substring heuristic
    cols_norm = {c: normalize_str(c) for c in df.columns}
    for c, c_norm in cols_norm.items():
        for cn in cand_norms:
            if cn and cn in c_norm:
                return c

    return None


def infer_columns(df: pd.DataFrame):
    col_brand = find_column(df, ["Brand", "brand"])
    col_acq = find_column(df, ["Acquisition", "acquisition", "Acq", "acq"])
    col_class = find_column(df, ["class", "Class", "botanical", "botanical_class", "honey_class"])
    col_conc = find_column(df, ["concentration", "Concentration", "sugar_concentration", "SugarConcentration"])
    col_conc_class = find_column(df, ["concentration_class", "Concentration_Class", "conc_class", "adulteration_class"])

    return {
        "Brand": col_brand,
        "Acquisition": col_acq,
        "class": col_class,
        "concentration": col_conc,
        "concentration_class": col_conc_class
    }


def infer_spectral_columns(df: pd.DataFrame, meta_cols: list) -> list:
    candidate_cols = [c for c in df.columns if c not in meta_cols]
    numeric_cols = [c for c in candidate_cols if pd.api.types.is_numeric_dtype(df[c])]
    return numeric_cols


def add_df_to_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame, title: str = None):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)

    if title:
        ws.cell(row=1, column=1, value=title).font = Font(bold=True)
        ws.append([])

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    header_row = 1 if not title else 3
    for cell in ws[header_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def save_confusion_matrix_figure(cm: np.ndarray, class_labels: list, out_path: str, title: str):
    fig = plt.figure(figsize=(8, 6))
    plt.imshow(cm, interpolation="nearest")
    plt.title(title)
    plt.xlabel("Predicted")
    plt.ylabel("True")
    plt.xticks(np.arange(len(class_labels)), class_labels, rotation=45, ha="right")
    plt.yticks(np.arange(len(class_labels)), class_labels)

    thresh = cm.max() / 2.0 if cm.max() > 0 else 0.5
    for i in range(cm.shape[0]):
        for j in range(cm.shape[1]):
            plt.text(
                j, i, f"{cm[i, j]}",
                ha="center", va="center",
                color="white" if cm[i, j] > thresh else "black"
            )

    plt.tight_layout()
    plt.savefig(out_path, dpi=300, bbox_inches="tight")
    plt.close(fig)


# -----------------------------
# MAIN
# -----------------------------
def main():
    setup_folders()
    logger = setup_logger()

    logger.info("Starting Random Forest multiclass classification (RandomizedSearchCV).")
    logger.info(f"Data path: {DATA_PATH}")
    logger.info(f"Outputs root: {STEP_DIR}")

    # ---- Load data
    df_raw = read_excel_dataset(DATA_PATH)
    logger.info(f"Loaded dataset with shape: {df_raw.shape}")

    # ---- Infer required columns
    cols = infer_columns(df_raw)
    logger.info("Resolved column mapping (internal -> actual):")
    for k, v in cols.items():
        logger.info(f"  {k:>20s} -> {v}")

    missing = [k for k, v in cols.items() if v is None]
    if missing:
        logger.error("Could not resolve required columns from the dataset.")
        logger.error(f"Missing internal fields: {missing}")
        logger.error("Available columns in Excel:")
        logger.error(df_raw.columns.tolist())
        raise ValueError(
            f"Missing required columns (not found in Excel): {missing}. "
            f"Check log for available column names: {LOG_FILE}"
        )

    # ---- Standardize internal names
    df = df_raw.rename(columns={
        cols["Brand"]: "Brand",
        cols["Acquisition"]: "Acquisition",
        cols["class"]: "class",
        cols["concentration"]: "concentration",
        cols["concentration_class"]: "concentration_class"
    }).copy()

    # ---- Split: Acquisition 1–5 vs 6
    train_df = df[df["Acquisition"].isin([1, 2, 3, 4, 5])].copy()
    test_df = df[df["Acquisition"] == 6].copy()

    if train_df.empty or test_df.empty:
        raise ValueError("Train or test split is empty. Check 'Acquisition' values in the dataset.")

    logger.info(f"Train set: {train_df.shape} (Acq 1–5)")
    logger.info(f"Test set:  {test_df.shape} (Acq 6)")

    # ---- Features (spectral bands only)
    meta_cols = ["Brand", "class", "Acquisition", "concentration", "concentration_class"]
    spectral_cols = infer_spectral_columns(df, meta_cols=meta_cols)

    if len(spectral_cols) < 10:
        logger.error("Too few inferred spectral columns. Dumping dtypes:")
        logger.error(df.dtypes)
        raise ValueError(
            f"Too few inferred spectral columns: {len(spectral_cols)}. "
            "Please verify which columns are the spectral bands."
        )

    logger.info(f"Inferred spectral feature columns: {len(spectral_cols)}")

    X_train = train_df[spectral_cols].to_numpy(dtype=float)
    y_train = train_df["concentration_class"].astype(str).to_numpy()

    X_test = test_df[spectral_cols].to_numpy(dtype=float)
    y_test = test_df["concentration_class"].astype(str).to_numpy()

    # ---- Diagnostic scaling check (train only) — no re-scaling
    band_means = np.mean(X_train, axis=0)
    band_stds = np.std(X_train, axis=0, ddof=0)

    scaling_summary = pd.DataFrame({
        "n_features": [len(spectral_cols)],
        "mean_of_means": [float(np.mean(band_means))],
        "sd_of_means": [float(np.std(band_means))],
        "mean_of_sds": [float(np.mean(band_stds))],
        "sd_of_sds": [float(np.std(band_stds))]
    })

    scaling_by_feature = pd.DataFrame({
        "feature": spectral_cols,
        "train_mean": band_means,
        "train_sd": band_stds
    })

    logger.info(
        "Scaling diagnostic (train): "
        f"mean_of_means={scaling_summary.loc[0, 'mean_of_means']:.4f}, "
        f"mean_of_sds={scaling_summary.loc[0, 'mean_of_sds']:.4f}"
    )

    # ---- Class counts in train vs test
    train_counts = train_df["concentration_class"].astype(str).value_counts().sort_index()
    test_counts = test_df["concentration_class"].astype(str).value_counts().sort_index()
    class_counts = pd.DataFrame({"train_count": train_counts, "test_count": test_counts}).fillna(0).astype(int)

    def sort_key(x):
        try:
            return float(x)
        except Exception:
            return x

    class_labels = sorted(class_counts.index.tolist(), key=sort_key)

    # ---- Base model
    rf = RandomForestClassifier(
        random_state=RANDOM_STATE,
        n_jobs=-1,
        class_weight="balanced"
    )

    # ---- Parameter distributions (sampled by RandomizedSearchCV)
    # We use small, practical ranges for laptop-friendly tuning.
    rng = np.random.RandomState(RANDOM_STATE)

    # RandomizedSearchCV accepts lists; we provide pre-generated samples for integer params
    n_estimators_list = rng.randint(300, 701, size=50).tolist()           # 300–700
    min_samples_split_list = rng.randint(2, 11, size=50).tolist()         # 2–10
    min_samples_leaf_list = rng.randint(1, 6, size=50).tolist()           # 1–5

    PARAM_DIST = {
        "n_estimators": sorted(list(set(n_estimators_list))),
        "max_depth": [None, 15, 30],
        "min_samples_split": sorted(list(set(min_samples_split_list))),
        "min_samples_leaf": sorted(list(set(min_samples_leaf_list))),
        "max_features": ["sqrt", "log2", 0.3]
    }

    # ---- CV + randomized tuning (training only)
    cv = StratifiedKFold(n_splits=N_SPLITS_CV, shuffle=True, random_state=RANDOM_STATE)

    search = RandomizedSearchCV(
        estimator=rf,
        param_distributions=PARAM_DIST,
        n_iter=N_ITER,
        scoring=SCORING,
        cv=cv,
        random_state=RANDOM_STATE,
        n_jobs=-1,
        refit=True,
        return_train_score=True
    )

    logger.info(f"Running RandomizedSearchCV (training only) with n_iter={N_ITER} ...")
    t0 = time.time()
    search.fit(X_train, y_train)
    elapsed = time.time() - t0

    logger.info(f"RandomizedSearchCV completed in {elapsed:.1f} s")
    logger.info(f"Best params: {search.best_params_}")
    logger.info(f"Best CV {SCORING}: {search.best_score_:.4f}")

    # ---- Tuning results table
    tune_results = pd.DataFrame(search.cv_results_).sort_values(by="rank_test_score")
    keep_cols = []
    for c in [
        "rank_test_score",
        "mean_test_score", "std_test_score",
        "mean_train_score", "std_train_score",
        "param_n_estimators", "param_max_depth",
        "param_min_samples_split", "param_min_samples_leaf",
        "param_max_features"
    ]:
        if c in tune_results.columns:
            keep_cols.append(c)

    tune_results_small = tune_results[keep_cols].copy()
    rename_map = {
        "mean_test_score": f"cv_mean_{SCORING}",
        "std_test_score": f"cv_sd_{SCORING}",
        "mean_train_score": f"train_mean_{SCORING}",
        "std_train_score": f"train_sd_{SCORING}",
        "param_n_estimators": "n_estimators",
        "param_max_depth": "max_depth",
        "param_min_samples_split": "min_samples_split",
        "param_min_samples_leaf": "min_samples_leaf",
        "param_max_features": "max_features"
    }
    tune_results_small.rename(columns={k: v for k, v in rename_map.items() if k in tune_results_small.columns}, inplace=True)

    best_params_df = pd.DataFrame([{
        "best_params_json": json.dumps(search.best_params_),
        f"best_cv_{SCORING}": float(search.best_score_),
        "cv_folds": N_SPLITS_CV,
        "n_iter": N_ITER,
        "random_state": RANDOM_STATE,
        "train_split": "Acquisition 1–5",
        "test_split": "Acquisition 6",
        "class_weight": "balanced",
        "model": "RandomForestClassifier",
        "tuning": "RandomizedSearchCV"
    }])

    # ---- External test evaluation
    best_model = search.best_estimator_
    y_pred = best_model.predict(X_test)

    acc = accuracy_score(y_test, y_pred)
    bal_acc = balanced_accuracy_score(y_test, y_pred)
    f1_macro = f1_score(y_test, y_pred, average="macro")

    logger.info(f"TEST accuracy:           {acc:.4f}")
    logger.info(f"TEST balanced accuracy:  {bal_acc:.4f}")
    logger.info(f"TEST macro-F1:           {f1_macro:.4f}")

    metrics_df = pd.DataFrame([{
        "test_accuracy": acc,
        "test_balanced_accuracy": bal_acc,
        "test_f1_macro": f1_macro
    }])

    report_dict = classification_report(y_test, y_pred, output_dict=True, zero_division=0)
    report_df = pd.DataFrame(report_dict).T.reset_index().rename(columns={"index": "label"})

    cm = confusion_matrix(y_test, y_pred, labels=class_labels)
    cm_df = pd.DataFrame(cm, index=[f"true_{c}" for c in class_labels], columns=[f"pred_{c}" for c in class_labels])

    save_confusion_matrix_figure(
        cm=cm,
        class_labels=class_labels,
        out_path=CONF_MATRIX_FIG,
        title="Random Forest — Confusion Matrix (External Test: Acquisition 6)"
    )
    logger.info(f"Saved confusion matrix figure: {CONF_MATRIX_FIG}")

    # ---- Feature importance (RF-specific; interpret cautiously)
    importances = getattr(best_model, "feature_importances_", None)
    fi_df = None
    if importances is not None:
        fi_df = pd.DataFrame({
            "feature": spectral_cols,
            "importance": importances
        }).sort_values(by="importance", ascending=False).reset_index(drop=True)

    # ---- Save to Excel
    wb = Workbook()
    wb.remove(wb.active)

    cols_df = pd.DataFrame({"excel_columns": df_raw.columns.tolist()})
    add_df_to_sheet(wb, "Excel_Columns", cols_df, title="Columns detected in the Excel file")

    mapping_df = pd.DataFrame([{"internal_name": k, "actual_excel_column": cols[k]} for k in cols.keys()])
    add_df_to_sheet(wb, "Column_Mapping", mapping_df, title="Resolved column mapping (internal -> Excel)")

    add_df_to_sheet(
        wb,
        "Data_Split",
        class_counts.reset_index().rename(columns={"index": "concentration_class"}),
        title="Class counts in Train (Acq 1–5) vs Test (Acq 6)"
    )

    add_df_to_sheet(wb, "Scaling_Summary", scaling_summary, title="Scaling diagnostic summary (train only)")
    add_df_to_sheet(wb, "Scaling_ByFeature", scaling_by_feature, title="Per-feature mean and SD (train only)")

    add_df_to_sheet(wb, "Tuning_Results", tune_results_small, title=f"RandomizedSearchCV results (scoring={SCORING})")
    add_df_to_sheet(wb, "Best_Params", best_params_df, title="Best hyperparameters and CV setup")

    add_df_to_sheet(wb, "Test_Metrics", metrics_df, title="External test performance (Acquisition 6)")
    add_df_to_sheet(
        wb,
        "Confusion_Matrix",
        cm_df.reset_index().rename(columns={"index": "true_label"}),
        title="Confusion matrix (external test)"
    )
    add_df_to_sheet(wb, "Class_Report", report_df, title="Classification report (external test)")

    if fi_df is not None:
        add_df_to_sheet(wb, "Feature_Importance", fi_df, title="Feature importance (Random Forest)")

    wb.save(EXCEL_OUT)
    logger.info(f"Saved Excel results: {EXCEL_OUT}")
    logger.info("Pipeline completed successfully.")


if __name__ == "__main__":
    main()
